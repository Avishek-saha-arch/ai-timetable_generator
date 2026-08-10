
#! RUN THIS FILE TO GENERATE THE ACTUAL TIMETABLE EXCEL FILE    
from __future__ import annotations
from pathlib import Path
import json
import pandas as pd
from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------
# SAMPLE JSON DATA STRUCTURE
# ---------------------------------------------------------
SAMPLE_JSON_INPUT = """
{
  "school_name": "ABC SCHOOL",
  "class_name": "Grade 10",
  "section": "A",
  "num_periods": 6,
  "period_duration_mins": 45,
  "days": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
  "subjects": [
    {"name": "Mathematics", "teachers": ["Mr. Sharma", "Prof. Alan"], "hours": 8},
    {"name": "Physics", "teachers": ["Dr. Robert"], "hours": 7},
    {"name": "Computer Science", "teachers": ["Prof. Sarah", "Prof. John"], "hours": 10},
    {"name": "Free Period / Self Study", "teachers": ["N/A"], "hours": 5}
  ]
}
"""

# ---------------------------------------------------------
# EXCEL GENERATOR
# ---------------------------------------------------------
def save_timetable_excel(
    rows: list,
    config: dict,
    filename: str = "generated_timetable.xlsx"
):
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="4472C4")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    cell_fill = PatternFill("solid", fgColor="DCE6F1")

    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=14)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    days = config["days"]
    periods = list(range(1, config["num_periods"] + 1))
    class_label = f"{config['class_name']} - Section {config['section']}"

    ws = wb.create_sheet(title=f"Class {config['section']}")

    # Title Header
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(days) + 1)
    title = ws["A1"]
    title.value = f"{config.get('school_name', 'SCHOOL')} | Timetable: {class_label} ({config['period_duration_mins']} mins/period)"
    title.fill = title_fill
    title.font = title_font
    title.alignment = center

    # Table Header Row
    ws["A2"] = "Period"
    ws["A2"].fill = header_fill
    ws["A2"].font = white_font
    ws["A2"].alignment = center

    for col, day in enumerate(days, start=2):
        cell = ws.cell(row=2, column=col)
        cell.value = day
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    # Period Column Labels
    for row_no, period in enumerate(periods, start=3):
        cell = ws.cell(row=row_no, column=1)
        cell.value = f"Period {period}"
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    # Populate Timetable Cells
    for day, period, subject, teacher in rows:
        row = period + 2
        col = days.index(day) + 2

        cell = ws.cell(row=row, column=col)
        cell.value = f"{subject}\n({teacher})"
        cell.alignment = center
        cell.fill = cell_fill
        cell.border = border

    ws.freeze_panes = "B3"

    # Layout Formatting
    ws.column_dimensions["A"].width = 16
    for c in range(2, len(days) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 24

    for r in range(3, len(periods) + 3):
        ws.row_dimensions[r].height = 45

    wb.save(filename)
    print(f"Excel timetable saved to {filename}")


# ---------------------------------------------------------
# OPTIMIZED OR-TOOLS SAT SOLVER
# ---------------------------------------------------------
def generate_timetable_from_json(json_input: str | dict, output_path: str = "generated_timetable.xlsx") -> pd.DataFrame:
    config = json.loads(json_input) if isinstance(json_input, str) else json_input

    days = config["days"]
    num_periods = config["num_periods"]
    periods = list(range(1, num_periods + 1))
    total_slots = len(days) * num_periods

    subjects_data = config["subjects"]
    total_requested_hours = sum(s["hours"] for s in subjects_data)

    # Capacity Check
    if total_requested_hours != total_slots:
        raise ValueError(
            f"Total requested hours ({total_requested_hours}) does not match available weekly slots ({total_slots})."
        )

    model = cp_model.CpModel()

    # Decision variables: x[(day, period, subject_idx)]
    x = {}
    for day in days:
        for period in periods:
            for s_idx, _ in enumerate(subjects_data):
                x[(day, period, s_idx)] = model.NewBoolVar(f"slot_{day}_{period}_s{s_idx}")

    # Constraint 1: Exactly 1 subject per time slot
    for day in days:
        for period in periods:
            model.Add(sum(x[(day, period, s_idx)] for s_idx, _ in enumerate(subjects_data)) == 1)

    # Constraint 2: Subject total weekly hours requirement
    for s_idx, s_info in enumerate(subjects_data):
        model.Add(
            sum(x[(day, period, s_idx)] for day in days for period in periods) == s_info["hours"]
        )

    # Constraint 3: Prevent same subject back-to-back in consecutive periods on same day
    for day in days:
        for period in periods[:-1]:
            for s_idx, _ in enumerate(subjects_data):
                model.Add(x[(day, period, s_idx)] + x[(day, period + 1, s_idx)] <= 1)

    # Solve
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 10
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError(f"Unable to find a valid timetable solution. Solver status: {status}")

    # Process Results
    results = []
    for day in days:
        for period in periods:
            for s_idx, s_info in enumerate(subjects_data):
                if solver.Value(x[(day, period, s_idx)]):
                    # Round-robin or primary teacher assignment
                    assigned_teacher = s_info["teachers"][0] if s_info["teachers"] else "N/A"
                    results.append([day, period, s_info["name"], assigned_teacher])

    df = pd.DataFrame(results, columns=["Day", "Period", "Subject", "Teacher"])

    # Export to Excel
    save_timetable_excel(results, config, filename=output_path)
    return df


if __name__ == "__main__":
    timetable_df = generate_timetable_from_json(SAMPLE_JSON_INPUT, "school_timetable.xlsx")
    print("\n--- Timetable Generated Successfully ---")
    print(timetable_df.to_string(index=False))
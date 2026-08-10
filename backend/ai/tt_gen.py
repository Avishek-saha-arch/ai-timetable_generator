from __future__ import annotations
import json
import io
import pandas as pd
from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def solve_timetable_grid(config: dict) -> dict:
    """Solves the constraint model and returns a JSON-friendly dict for React frontend.
    
    Output structure:
    {
        "Monday": [
            {"subject": "Mathematics", "teacher": "Mr. Smith", "type": "Lecture"},
            ...
        ],
        ...
    }
    """
    days = config.get("days", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"])
    num_periods = config.get("numPeriods", config.get("num_periods", 6))
    periods = list(range(0, num_periods))  # ZERO-BASED for JS array alignment
    total_slots = len(days) * num_periods

    subjects_data = config.get("subjects", [])
    total_requested_hours = sum(int(s.get("hours", 0)) for s in subjects_data)

    if total_requested_hours != total_slots:
        raise ValueError(
            f"Total requested hours ({total_requested_hours}) does not match total available weekly slots ({total_slots})."
        )

    model = cp_model.CpModel()
    x = {}

    for day in days:
        for period in periods:
            for s_idx, _ in enumerate(subjects_data):
                x[(day, period, s_idx)] = model.NewBoolVar(f"slot_{day}_{period}_s{s_idx}")

    # Constraint 1: Exactly one subject per slot
    for day in days:
        for period in periods:
            model.Add(sum(x[(day, period, s_idx)] for s_idx, _ in enumerate(subjects_data)) == 1)

    # Constraint 2: Subject total weekly hours
    for s_idx, s_info in enumerate(subjects_data):
        model.Add(
            sum(x[(day, period, s_idx)] for day in days for period in periods) == int(s_info.get("hours", 0))
        )

    # Constraint 3: No back-to-back same subjects
    for day in days:
        for period in periods[:-1]:
            for s_idx, _ in enumerate(subjects_data):
                model.Add(x[(day, period, s_idx)] + x[(day, period + 1, s_idx)] <= 1)

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 10
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError(f"Unable to generate timetable. Solver status: {status}")

    # Build the normalized dictionary output structured by Day -> Period index
    schedule = {day: [None] * num_periods for day in days}

    for day in days:
        for period in periods:
            for s_idx, s_info in enumerate(subjects_data):
                if solver.Value(x[(day, period, s_idx)]):
                    teachers = s_info.get("teachers", [])
                    # Join teachers list if array provided
                    teacher_str = ", ".join([t for t in teachers if t]) if isinstance(teachers, list) else str(teachers)
                    
                    is_free = "Free" in s_info.get("name", "") or s_info.get("id") == "free_period_row"

                    schedule[day][period] = {
                        "subject": s_info.get("name", "Unassigned"),
                        "teacher": teacher_str if teacher_str else "N/A",
                        "type": "Break" if is_free else "Lecture",
                        "color": "emerald" if is_free else "blue",
                        "room": f"Room {101 + s_idx}"
                    }

    return schedule


def save_timetable_excel_stream(rows: list, config: dict) -> io.BytesIO:
    """Generates the openpyxl workbook directly into an in-memory BytesIO buffer."""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="4472C4")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    cell_fill = PatternFill("solid", fgColor="DCE6F1")

    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=14)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    days = config.get("days", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"])
    num_periods = config.get("num_periods", 6)
    periods = list(range(1, num_periods + 1))
    class_label = f"{config.get('className', 'Class')} - Section {config.get('section', 'A')}"

    ws = wb.create_sheet(title=f"Class {config.get('section', 'A')}")

    # Header title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(days) + 1)
    title = ws["A1"]
    title.value = f"{config.get('school_name', 'SCHOOL')} | Timetable: {class_label} ({config.get('periodDuration', 45)} mins/period)"
    title.fill = title_fill
    title.font = title_font
    title.alignment = center

    # Table headers
    ws["A2"] = "Period"
    ws["A2"].fill = header_fill
    ws["A2"].font = white_font
    ws["A2"].alignment = center

    for col, day in enumerate(days, start=2):
        cell = ws.cell(row=2, column=col)
        cell.value = day
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    # Period column
    for row_no, period in enumerate(periods, start=3):
        cell = ws.cell(row=row_no, column=1)
        cell.value = f"Period {period}"
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    # Fill cell entries
    for day, period, subject, teacher in rows:
        row = period + 2
        col = days.index(day) + 2

        cell = ws.cell(row=row, column=col)
        cell.value = f"{subject}\n({teacher})"
        cell.alignment = center
        cell.fill = cell_fill
        cell.border = border

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 16

    for c in range(2, len(days) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 24

    for r in range(3, len(periods) + 3):
        ws.row_dimensions[r].height = 45

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer